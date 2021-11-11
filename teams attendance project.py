# Online Python compiler (interpreter) to run Python online.
# Write Python 3 code in this online editor and run it.
#F I N A L  C O P Y 

import openpyxl
import os
from openpyxl.styles import Font
from tkinter import *
import csv
from openpyxl import Workbook
from tkinter import filedialog
def calc(sheet,sheet2,sheet1):
 r=sheet.max_row
 #c=sheet.max_column
 r2=sheet2.max_row
 c2=sheet2.max_column
 nmls1=[]
 nmls2=[]
 for i in range(2,r2+1):
   k=(sheet2.cell(row=i,column=4).value).split(" ")
   print("k ::::: ",k)
   nmls2.append((sheet2.cell(row=i,column=4).value).split(" "))  # THE NAME ALGORITHM
   pos=[]
   for j in range(len(k)):
     #print("len(k) ",len(k))
     n1=len(k[j])
     if(n1 <= 2):
       pos.append(k[j]) 
     #print("pos ",pos) 
   for m in range(len(pos)):
     #print("inside")
     #print("POS[M] is......",pos[m])
     #print("List K before deleting......",k)
     k.remove(pos[m])
     #print("List K AFTERRR deleting......",k)    
   nmls1.append(k)
 print("nmls2 \n",nmls2)
 print("\nnmls1 ::: \n",nmls1)

 #print("total columns ",c2)
 #dy1,mth1,yr1=dt.split(":")
 ct=0
 ct2=0
 ai,ti=sheet.cell(row=3,column=3).value.split(",") # row=3 cuz even if maam joins way before time it checks either with their in between leaving time or next students joining time.
 mth1,dy1,yr1=ai.split("/")
 dtls=[]
 #print("ti == ",ti)
 hi,mi,si=ti.split(":")
 h4=0
 m4=0
  #print("hi, mi ====",hi,"  ",mi)
 tmls=[[[8,20],[9,30]],[[9,31],[10,45]],[[10,46],[11,59]],[[12,30],[1,59]],[[2,5],[3,15]]]
 for i in range(len(tmls)):
    #print("tmls[i] ",tmls[i],"   tmls[i][0][0] ",tmls[i][0][0]) 
    if((int(hi)==tmls[i][0][0] and int(mi)>=tmls[i][0][1]) or (int(hi)==tmls[i][1][0] and int(mi)<tmls[i][1][1])):
      #print("**inside**")
      h4=tmls[i][1][0]
      m4=tmls[i][1][1]
      #print("END TIME OF CLASS: ",h4,":",m4)
      break  
 for i in range(5,c2+1):
   temp=[]
   if(c2>4): 
    q,f=str(sheet2.cell(row=1,column=i).value).split(',')  ########
    #print("old date: ",q)
    mth2,dy2,yr2=q.split("/")
    temp.append(dy2)
    temp.append(mth2)
    temp.append(yr2)
    dtls.append(temp)
    if(dy1==dy2 and mth1==mth2 and yr1==yr2):
      ct=ct+1
    f2,f3=f.split('-')
    f4,fm4=f3.split(':')
    #print('f4   m4   === ',f4,'  ',m4)
    if(int(f4)==h4):
      ct2=ct2+1
 
 if(ct==0 or ct2==0):
  
  #print("****")
  l1=[]  
  
  i=2
  ls=[]
  while(i<r+1):
    j=i
    tm=0
    l=[]  
    a,t1=sheet.cell(row=j,column=3).value.split(",")
    h1,m1,s1=t1.split(":")
    if(sheet.cell(row=j+1,column=2).value=="Left"):
       a,t2=sheet.cell(row=j+1,column=3).value.split(",")
       h2,m2,s2=t2.split(":")
       tm=(((int(h2)-int(h1))*60)+(int(m2)-int(m1)))
       i=i+2     	  
    elif((sheet.cell(row=j+1,column=1).value == sheet.cell(row=j,column=1).value) and sheet.cell(row=j+1,column=2).value=="Joined"):
       a,t2=sheet.cell(row=j+1,column=3).value.split(",")
       h2,m2,s2=t2.split(":")
       tm=(((int(h2)-int(h1))*60)+(int(m2)-int(m1)))
       i=i+1
    else:
      tm=(((int(h4)-int(h1))*60)+(int(m4)-int(m1)))
      i=i+1 
    if(i>3 and sheet.cell(row=j,column=1).value == sheet.cell(row=j-1,column=1).value and (sheet.cell(row=j-1,column=2).value=="Left" or sheet.cell(row=j-1,column=2).value=="Joined") ):  #For repeating cases of names
      a_temp,t_temp=sheet.cell(row=j-1,column=3).value.split(",")
      h_temp,m_temp,s_temp=t_temp.split(":")
      if(int(h1)>=int(h_temp) or int(m1)>=int(m_temp)): 
       ls[len(ls)-1][1]=ls[len(ls)-1][1]+tm

    else:
       l.append(sheet.cell(row=j,column=1).value)
       l.append(tm)
       ls.append(l)
       #print(ls) 
  l_chk=[]
  #print("list printing...",ls)
  for i in range(len(ls)):
    k=(ls[i][0].upper()).split(" ")
    l_chk.append((ls[i][0].upper()).split(" "))
    pos=[]
    for j in range(len(k)):
      if((len(k[j]))<2):
        pos.append(k[j])
      #print("pos ",pos)  
    for m in range(len(pos)):
        #print("inside")
        #print("POS[M] is......",pos[m])
        #print("List K before deleting......",k)
        k.remove(pos[m])
        #print("List K before deleting......",k)
		    
    ls[i][0]=k
    #print("ls[i][0] ",ls[i][0])
	  
      
  temp1=[]
  temp2=[]
  temp3=[]
  for i in range(len(dtls)):
    if(int(yr1)<int(dtls[i][2])):	
      temp1.append(dtls[i])
  for i in range(len(dtls)):
    if(int(mth1)<int(dtls[i][1]) and int(yr1)==int(dtls[i][2])):
      temp2.append(dtls[i])
  for i in range(len(dtls)):
    if(int(dy1)<int(dtls[i][0]) and int(mth1)==int(dtls[i][1]) and int(yr1)==int(dtls[i][2])):
      temp3.append(dtls[i])
  for i in range(len(temp1)):
    dtls.remove(temp1[i])
  for i in range(len(temp2)):
    dtls.remove(temp2[i])	
  for i in range(len(temp3)):
    dtls.remove(temp3[i])


  #print("len(dtls) later : ",len(dtls))
  n=len(dtls)+4
  if(c2>4 and n<c2):
    dt_new,timing=(sheet2.cell(row=1,column=n+1).value).split(',')
  else:
    dt_new=sheet2.cell(row=1,column=n+1).value
  sheet2.insert_cols(idx=n+1)    #creating column 1   
  #sheet.insert_cols(idx=c+2,amount=r)    #creating column 2
  d={}
  for i in range(len(ls)):
    ld=[]
    if(ls[i][1]>=40):  #criteria : 40 minutes as the qualifying time for PRESENT
      ls[i].append("P")
      #print('\n',ls[i][0]," - ",ls[i][2])
      #ld.append(ls[i][1])
      ld.append(ls[i][0])
      ld.append(ls[i][2])
      ld.append(ls[i][1])
      d[str(ls[i][0])]=ld

    #elif(ls[i][1]==-1):
      #ls[i].append("PRESENT")
      #print('\n',ls[i][0],' - ',ls[i][2])
      #ld.append("FULL TIME")
      #ld.append(ls[i][2])
      #d[ls[i][0]]=ld	   	       	 	    

   
    else: 
      ls[i].append("Ab")
      #print('\n',ls[i][0]," - ",ls[i][2])
      #d[ls[i][0]]=ls[i][2]
      #ld.append(ls[i][1])
      
      ld.append(ls[i][0])
      ld.append(ls[i][2])
      ld.append(ls[i][1])
      d[str(ls[i][0])]=ld
  #print("check***",ls)   	 
  #sheet.cell(row=1,column=c+1).value="Time(minutes)"
  sheet2.cell(row=1,column=n+1).value=ai+','+str(h4-1)+':'+str(m4)+'-'+str(h4)+':'+str(m4)
  #n=sheet.cell(row=3,column=1).value
  #sheet2.cell(row=(l1.index(d_org[d[n][0]])+1),column=c+1).value=d[n][1]
  #sheet.cell(row=3,column=c+2).value=d[n][1]
  #print("l_chk : ",l_chk)
  #print("\n nmls2 : ",nmls2)
  time_ls=[]
  for i in range(len(ls)):
   tmp=[]  
   if(ls[i][0] in nmls1):
    #u=l1.index(d_org[d[sheet.cell(row=i,column=1).value][0]])+1
    if(nmls1.count(ls[i][0])==1):  
      u=nmls1.index(ls[i][0])
      sheet2.cell(row=u+2,column=n+1).value = d[str(ls[i][0])][1]
      tmp.append(sheet2.cell(row=u+2,column=4).value)
      tmp.append(d[str(ls[i][0])][2])
      time_ls.append(tmp)
      #print("print 1 ::: ",tmp)
    else:
      for k in range(nmls1.count(ls[i][0])):
        t_pos=(nmls1.index(ls[i][0]))+k
        #print("ls[i][0] :::: ",ls[i][0])
        #print("nmls1.count(ls[i][0]) ::::  ",nmls1.count(ls[i][0]))
        
        #print("t_pos ::: ",t_pos)
        g=0
        for j in range((len(l_chk[i]))):
          if(l_chk[i][j] in nmls2[t_pos] and sheet2.cell(row=t_pos+2,column=n+1).value != 'P'):
            #print("nmls2[tpos] ::: ",nmls2[t_pos])
            g=g+1
        if(g==len(l_chk[i])):
          sheet2.cell(row=t_pos+2,column=n+1).value = d[str(ls[i][0])][1]
          tmp.append(sheet2.cell(row=t_pos+2,column=4).value)
          tmp.append(d[str(ls[i][0])][2])
          time_ls.append(tmp)
          #print("print 2 ::: ",tmp)
          
          #print("str(ls[i][0]) ::: ",str(ls[i][0]))
          #print("sheet2.cell(row=t_pos+2,column=n+1).value :::: ",sheet2.cell(row=t_pos+2,column=n+1).value)

          break



  abs1=[]  
  abs2=[]
  for i in range(2,r2+1):
    tmp=[]  
    if(sheet2.cell(row=i,column=n+1).value!="P" ):
      #f=sheet2.cell(row=i,column=c2+1)
      if(sheet2.cell(row=i,column=n+1).value != 'Ab'):
          tmp.append(sheet2.cell(row=i,column=4).value)
          tmp.append(0)
          time_ls.append(tmp)
          #print("print 3 ::: ",tmp)
          
          
      sheet2.cell(row=i,column=n+1).value="Ab"
      abs1.append(sheet2.cell(row=i,column=2).value)
      abs2.append(sheet2.cell(row=i,column=4).value)
      sheet2.cell(row=i,column=n+1).font = Font(color="00FF0000")
    else:
	    sheet2.cell(row=i,column=n+1).font = Font(color="0000FF00") 
      
    #sheet.cell(row=i,column=c+2).value = d[sheet.cell(row=i,column=1).value][1]              
  #nm=input("\nEnter name for the new updated excel file (ex: abc.xlsx ) : ") #saving in a new excel sheet
  ## CREATING A NEW TEXT FILE NAMED  "class_att_list_IT_A.txt"  TO SAVE IMMIDEATE DETAILS
  r1=sheet1.max_row
  c1=sheet2.max_column

  if(n!=c2):
    for i in range(1,r1+1):
      #print("inside1***")
      #print("dt_new :::: ",dt_new)
      print("sheet1.cell(row=i,column=1).value, dt_new  ",sheet1.cell(row=i,column=1).value,"  ;;;;  ",dt_new)
      if(sheet1.cell(row=i,column=1).value==dt_new):  
        #print("inside2***")
        print("sheet1.cell(row=i,column=1).value ::: ",sheet1.cell(row=i,column=1).value)
        sheet1.insert_rows(idx=i,amount=len(abs2)+2)
        r1=i-1
        #print("r1 (1) :::: ",r1)
        break


  if(n==c2 and r1!=1):
    r1=r1+2
    #print("r1 (2) ::: ",r1)
  
  sheet1.cell(row=r1+1,column=1).value=ai

  sheet1.cell(row=r1+1,column=2).value=str(h4-1)+":"+str(m4)+" - "+str(h4)+":"+str(m4)

  sheet1.cell(row=r1+1,column=3).value=str(r2-1-len(abs1))+" / "+str(len(abs1))

  #print("print 4 ::: ",time_ls)
  for i in range(len(abs1)):
    a=str(abs1[i])
    b=str(abs2[i])
    sheet1.cell(row=r1+i+1,column=4).value=a
    sheet1.cell(row=r1+i+1,column=5).value=b
    for j in range(len(time_ls)):
        if(abs2[i] == time_ls[j][0]):
          sheet1.cell(row=r1+i+1,column=6).value=time_ls[j][1]

  	
    
  
  wb2.save(filename = "class_list_IT_A.xlsx")
  print("\n\n   *** ATTENDANCE SUCCESSFULLY CALCULATED *** ")
  print("\n    Check class_list_IT_A.xlsx for details of compiled attendance")

 else:
   print("\n\n\n   Attendance for the date : ",ai," has already been compiled ")
   print("     Please check class_list_IT_A.xlsx for the compiled details....")


# OUTSIDE THE FUNCTION
print("\n     P R A J E E T H   T E A M S   A T T E N D A N C E    C A L C U L A T O R    ")
print("\n S E L E C T    F I L E ")
root=Tk()
root.withdraw()
root.name=filedialog.askopenfilename()


#ft=str(input("enter the end time of class in hh:mm format : "))
#h4,m4=ft.split(":")
#d_org={"B PRAJEETH":205002062} #example prototype yet to be completed

s=root.name
if(s[-4::1]=='.csv'):
 s=os.path.basename(root.name)
 wb = openpyxl.Workbook()
 ws = wb.active
 with open(root.name) as csv_file:
    csv_reader = csv.reader((line.replace('\0','') for line in csv_file),delimiter='\t')
    for row in csv_reader:
        ws.append(row)
		
 #print("\n\n\nroot.name : ",root.name)
 print(" s= ",s)
 s=s.replace(".csv",'')
 s=s+'.xlsx'
 wb.save(s)		
 os.remove(root.name)

wb = openpyxl.load_workbook(s)

wb2= openpyxl.load_workbook("class_list_IT_A.xlsx")
x=wb.get_sheet_names()
y=wb2.get_sheet_names()
sheet = wb.get_sheet_by_name(x[0])
sheet1=wb2.get_sheet_by_name(y[1])
sheet2= wb2.get_sheet_by_name(y[0])

calc(sheet,sheet2,sheet1)
root.mainloop()
