#after including dialog Box
import openpyxl
from openpyxl.styles import Font
from tkinter import *
from tkinter import filedialog
def calc(sheet,sheet2,d_org,dt):
 r=sheet.max_row
 #c=sheet.max_column
 r2=sheet2.max_row
 c2=sheet2.max_column
 print("total columns ",c2)
 dy1,mth1,yr1=dt.split(":")
 ct=0
 dtls=[]
 for i in range(5,c2+1):
   temp=[]
   if(c2>4): 
    q=str(sheet2.cell(row=1,column=i).value)
    print("old date: ",q)
    dy2,mth2,yr2=q.split(":")
    temp.append(dy2)
    temp.append(mth2)
    temp.append(yr2)
    dtls.append(temp)
    if(dy1==dy2 and mth1==mth2 and yr1==yr2):
      ct=ct+1
 if(ct==0):
  
  #print("****")
  l1=[]  
  for i in range(0,r2):
    l1.append(sheet2.cell(row=i+1,column=2).value)


  
  print("dtls  ",dtls)
  #print("####")
  ls=[]
  i=2
  while(i<r+1):
    j=i
    tm=0
    l=[]  
    a,t1=sheet.cell(row=j,column=3).value.split(",")
    h1,m1,s1=t1.split(":")
    if(sheet.cell(row=j+1,column=2).value=="Left"):
       a,t2=sheet.cell(row=j+1,column=3).value.split(",")
       h2,m2,s2=t2.split(":")
       tm=(((int(h2)-int(h1))*60)+(int(m2)-int(m1)))
       i=i+2	     	  
    else:
      tm=(((int(h4)-int(h1))*60)+(int(m4)-int(m1)))
      i=i+1 
    if(i>3 and sheet.cell(row=j,column=1).value == sheet.cell(row=j-2,column=1).value and sheet.cell(row=j-2,column=2).value=="Joined" ):  #For repeating cases of names
       ls[len(ls)-1][1]=ls[len(ls)-1][1]+tm
    else:
       l.append(sheet.cell(row=j,column=1).value)
       l.append(tm)
       ls.append(l)
       #print(ls)   
	    
  for i in range(len(dtls)):
    if(int(yr1)<int(dtls[i][2])):
      dtls.remove(dtls[i])
  for i in range(len(dtls)):
    if(int(mth1)<int(dtls[i][1]) and int(yr1)==int(dtls[i][2])):
      dtls.remove(dtls[i])
  for i in range(len(dtls)):
    if(int(dy1)<int(dtls[i][0]) and int(mth1)==int(dtls[i][1]) and int(yr1)==int(dtls[i][2])):
      dtls.remove(dtls[i])
	  


  print("len(dtls) later : ",len(dtls))
  n=len(dtls)+4
  sheet2.insert_cols(idx=n+1)    #creating column 1   
  #sheet.insert_cols(idx=c+2,amount=r)    #creating column 2
  d={}
  for i in range(len(ls)):
    ld=[]
    if(ls[i][1]>50):  #criteria : 50 minutes as the qualifying time for PRESENT
      ls[i].append("P")
      print('\n',ls[i][0]," - ",ls[i][2])
      #ld.append(ls[i][1])
      ld.append(ls[i][0])
      ld.append(ls[i][2])
      d[ls[i][0]]=ld

    #elif(ls[i][1]==-1):
      #ls[i].append("PRESENT")
      #print('\n',ls[i][0],' - ',ls[i][2])
      #ld.append("FULL TIME")
      #ld.append(ls[i][2])
      #d[ls[i][0]]=ld	   	       	 	    

   
    else: 
      ls[i].append("Ab")
      print('\n',ls[i][0]," - ",ls[i][2])
      d[ls[i][0]]=ls[i][2]
      #ld.append(ls[i][1])
      ld.append(ls[i][0])
      ld.append(ls[i][2])
      d[ls[i][0]]=ld
  #print("check***",ls)   	 
  #sheet.cell(row=1,column=c+1).value="Time(minutes)"
  sheet2.cell(row=1,column=n+1).value=dt
  #n=sheet.cell(row=3,column=1).value
  #sheet2.cell(row=(l1.index(d_org[d[n][0]])+1),column=c+1).value=d[n][1]
  #sheet.cell(row=3,column=c+2).value=d[n][1]

  for i in range(2,r+1):
   if(sheet.cell(row=i-1,column=1).value != sheet.cell(row=i,column=1).value and (d[sheet.cell(row=i,column=1).value][0] in d_org )):   
    u=l1.index(d_org[d[sheet.cell(row=i,column=1).value][0]])+1
	  
    sheet2.cell(row=u,column=n+1).value = d[sheet.cell(row=i,column=1).value][1]
  abs1=[]  
  abs2=[]
  for i in range(2,r2+1):
    if(sheet2.cell(row=i,column=n+1).value!="P"):
      #f=sheet2.cell(row=i,column=c2+1)
      sheet2.cell(row=i,column=n+1).value="Ab"
      abs1.append(sheet2.cell(row=i,column=2).value)
      abs2.append(sheet2.cell(row=i,column=4).value)
      sheet2.cell(row=i,column=n+1).font = Font(color="00FF0000")
    else:
	    sheet2.cell(row=i,column=c2+1).font = Font(color="0000FF00") 
    #sheet.cell(row=i,column=c+2).value = d[sheet.cell(row=i,column=1).value][1]              
  #nm=input("\nEnter name for the new updated excel file (ex: abc.xlsx ) : ") #saving in a new excel sheet
  f=open("class_att_list_IT_A.txt","a+")
  f.write("date : "+str(dt))
  f.write("\n")
  f.write("total absent : "+str(len(abs1)))
  f.write("\n")
  f.write("total present : "+str(r2-1-len(abs1)))
  f.write("\n")
  f.write("<<< A B S E N T E E S >>>"+"\n")
  for i in range(len(abs1)):
    a=str(abs1[i])
    b=str(abs2[i])
    f.write(a+" - "+b)
    f.write("\n")
  f.write("\n")
  f.write("\n")  
  f.close  

  wb2.save(filename = "class_list_IT_A.xlsx")
 else:
   print("   Attendance for the date : ",dt," has already been compiled... ") 

# OUTSIDE THE FUNCTION
print("\n     P R A J E E T H   T E A M S   A T T E N D A N C E    C A L C U L A T O R    ")
print("\n S E L E C T    F I L E ")
root=Tk()
root.withdraw()
root.name=filedialog.askopenfilename()

#name=input("Enter the full location of the excel file : ")
ft=str(input("enter the end time of class in hh:mm format : "))
l=1
while(l==1):
    dt=input("enter date of class you uploaded in (dd:mm:yy) format ex:(05:05:21): ")
    if(dt.count(":")==2):
	      l=0
    else:
        print("***Not in format(dd:mm:yy)*** \n Please enter input again...")		
	

h4,m4=ft.split(":")
d_org={"B PRAJEETH":205002062,"Deepak":} #example prototype yet to be completed
wb = openpyxl.load_workbook(root.name)

wb2= openpyxl.load_workbook("class_list_IT_A.xlsx")
x=wb.get_sheet_names()
y=wb2.get_sheet_names()
sheet = wb.get_sheet_by_name(x[0])
sheet2= wb2.get_sheet_by_name(y[0])

calc(sheet,sheet2,d_org,dt)
root.mainloop()